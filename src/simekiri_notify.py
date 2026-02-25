#simekiri_notify.py

import os
import sys
import json
import traceback
from datetime import datetime, timedelta

import pandas as pd
import requests
from PIL import Image, ImageDraw, ImageFont
from io import BytesIO
import textwrap

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

LOG_FILE = None

def write_log(message):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ファイル出力
    if LOG_FILE:
        try:
            with open(LOG_FILE, "a", encoding="utf-8") as f:
                f.write(f"[{now}] {message}\n")
        except Exception:
            pass

    # コンソール出力
    print(f"[{now}] {message}")

def run_notify(config_path=None, test_mode=False):
    def pt_to_px(pt):
        return int(pt * 96 / 72)
    # ===== 基本パス =====
    if getattr(sys, 'frozen', False):
        BASE_DIR = os.path.dirname(sys.executable)
    else:
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))

    APP_DIR = os.path.join(os.environ["LOCALAPPDATA"], "SimekiriKyokan")
    os.makedirs(APP_DIR, exist_ok=True)
    
    global LOG_FILE
    LOG_FILE = os.path.join(APP_DIR, "simekiri_run_log.txt")

    # ---- args ----？？？らしい
    print("ARGV:", sys.argv)
    CONFIG_PATH = config_path
    print("CONFIG_PATH:", CONFIG_PATH)
    
    # config読み
    if not config_path:
        print("ERROR: config_path is None")
        return 1
    
    CONFIG_FILE = config_path
    write_log(f"Using config file: {CONFIG_FILE}")
    
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        config = json.load(f)
    
    # Webhook読み
    WEBHOOK_URL = config.get("webhook_url")
    if not WEBHOOK_URL:
        write_log("Webhook URL is missing")
        return 1
    
    # 通知設定読み
    DAYS_BEFORE = config.get("days_before_deadline", 3)
    MENTION_ENABLED = config.get("mention_enabled", False)
    MENTION_MAP = config.get("mentions", {})
    
    # list → dict
    if isinstance(MENTION_MAP, list):
        write_log("mentions is list → converting to dict")
    
        fixed = {}
        for item in MENTION_MAP:
            if not isinstance(item, dict):
                continue
    
            name = str(item.get("name", "")).strip()
            user_id = str(item.get("id", "")).strip()
    
            if name and user_id:
                fixed[name] = f"<@{user_id}>"
    
        MENTION_MAP = fixed
    
    write_log(f"MENTION_ENABLED={MENTION_ENABLED}")
    write_log(f"MENTION_MAP={MENTION_MAP}")
    
    # テスト通知実行
    IS_TEST = test_mode
    
    if IS_TEST:
        write_log("=== TEST MODE ===")
    
        try:
            r = requests.post(
                WEBHOOK_URL,
                json={
                    "content": "🧪 **締切教官 通知テスト**\nこのメッセージが見えていれば正常です。"
                }
            )
    
            status = getattr(r, "status_code", None)
            write_log(f"Test notify sent. status={status}")
    
            if status in (200, 204):
                return 0
            else:
                return 1
    
        except Exception as e:
            write_log("Test notify failed: " + repr(e))
            return 1
        
    # Paths and logging helpers
    # ↑？？？↓
    # fallback log dir: try BASE_DIR, if not writable use LOCALAPPDATA
    def _choose_log_dir():
        try:
            os.makedirs(APP_DIR, exist_ok=True)
            return APP_DIR
        except Exception:
            return os.path.expanduser("~")

    
    LOG_DIR = _choose_log_dir()
    LOG_FILE = os.path.join(LOG_DIR, "simekiri_run_log.txt")
    ERR_FILE = os.path.join(LOG_DIR, "simekiri_error_log.txt")

    write_log("=== start run pid=" + str(os.getpid()) + " cwd=" + os.getcwd() + " ===")
    
    # -------------------------
    # Safe main wrapper
    # -------------------------
    try:
        # -------------------------
        # Load config
        # -------------------------
        
        APP_DIR = os.path.join(os.environ["LOCALAPPDATA"], "SimekiriKyokan")
        os.makedirs(APP_DIR, exist_ok=True)
        
        EXCEL_FILE = config.get("excel_path")
        
        if not EXCEL_FILE or not os.path.exists(EXCEL_FILE):
            write_log(f"Excel NOT FOUND: {EXCEL_FILE}")
            return 1
        
        WEBHOOK_URL = config.get("webhook_url", "")
        DAYS_BEFORE_DEADLINE = config.get("days_before_deadline", 3)
        MENTION_ENABLED = config.get("mention_enabled", False)
        
        if not os.path.exists(EXCEL_FILE):
            write_log(f"Excel NOT FOUND: {EXCEL_FILE}")
            return 1
    
        # -------------------------
        # Helpers: date conversions
        # -------------------------
        def convert_deadline_value(x):
            # Accept Excel serial numbers, datetimes, strings
            if pd.isna(x):
                return pd.NaT
            # Excel serial (number)
            if isinstance(x, (int, float)):
                try:
                    # Excel's day 0 is 1899-12-30 for pandas compatibility
                    return (pd.to_datetime("1899-12-30") + pd.to_timedelta(x, unit="D"))
                except Exception:
                    return pd.to_datetime(x, errors="coerce")
            try:
                return pd.to_datetime(x, errors="coerce")
            except Exception:
                return pd.NaT
    
        # エクセル読み / 画像生成
        try:
            df = pd.read_excel(
                EXCEL_FILE,
                sheet_name="作業リスト",
                usecols="C:K"
            )

            # Excelの列幅取得（B列スタートで同期）
            wb = load_workbook(EXCEL_FILE)
            ws = wb["作業リスト"]
            
            excel_width_map = {}
            
            start_col_index = 3  # ← C列から開始って意味（A=1, B=2）
            
            for i, col_name in enumerate(df.columns):
                excel_col_index = start_col_index + i
                letter = get_column_letter(excel_col_index)
                dim = ws.column_dimensions.get(letter)
            
                if dim and dim.width:
                    # Excel列幅 → ピクセル換算（経験則係数）←？？？
                    pixel_width = int(dim.width * 8.2 + 12)
                    excel_width_map[col_name] = pixel_width
                else:
                    # 幅未設定時のフォールバック
                    excel_width_map[col_name] = 120
                    
            COL_WIDTH_MAP = excel_width_map

            write_log("COLUMNS: " + str(list(df.columns)))

            # 必須列チェック（Task化前の前提）下記の項目が未入力では通知されない
            REQUIRED_COLUMNS = ["内容", "締切", "担当", "進捗"]
    
            COLUMN_ORDER = list(df.columns)
            
            data_row_index = 8  # 実データ開始行に合わせて
            excel_row_height = ws.row_dimensions[data_row_index].height
            
            if excel_row_height:
                row_height_base = int(excel_row_height * 96 / 72) + 3
            else:
                row_height_base = pt_to_px(15)

            missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
            if missing:
                write_log(f"Missing columns: {missing}")
                return 1
                
            # 担当が空の行は除外
            df = df[df["担当"].notna()]
            df = df[df["担当"].astype(str).str.strip() != ""]
    
        except Exception as e:
            write_log("Failed to read excel: " + repr(e))
            raise
    
        # いらない列を除外（空白列だの）
        df.columns = df.columns.str.strip()
        df = df.drop(columns=[c for c in df.columns if "Unnamed" in c or c == "目次"], errors="ignore")
    
        def convert_done(v):
            return str(v).strip() == "完了"

        df["進捗_raw"] = df["進捗"]  # 元文字列保持
        df["進捗"] = df["進捗"].apply(convert_done)
	
        # normalize 締切　←？
        df["締切"] = df["締切"].apply(convert_deadline_value)
        df = df.dropna(subset=["内容", "締切"])
    
    
        # ensure datetime dtype　← 多分日付
        if not pd.api.types.is_datetime64_any_dtype(df["締切"]):
            df["締切"] = pd.to_datetime(df["締切"], errors="coerce")
    
        today = datetime.now().date()
        today_str = today.strftime("%Y%m%d")
    
        # compute rates
        total_tasks = len(df)
        completed_tasks = int(df["進捗"].sum()) if "進捗" in df else 0
        overall_rate = round((completed_tasks / total_tasks) * 100) if total_tasks > 0 else 0
    
        person_rates = {}
        if "担当" in df.columns:
            for person, g in df[df["担当"].notna()].groupby("担当"):
                total = len(g)
                done = int(g["進捗"].sum())
                person_rates[str(person).strip()] = round((done / total) * 100) if total > 0 else 0
    
        df["days_left"] = (df["締切"].dt.date - today)
        df["days_left"] = df["days_left"].apply(lambda x: x.days if pd.notna(x) else 9999)
        
        if "優先度" in df.columns:
            not_unnecessary = ~df["優先度"].astype(str).str.strip().isin(["不要"])
        else:
            not_unnecessary = True
        
        pending = df[
            (df["進捗"] == False) &
            not_unnecessary &
            (df["days_left"] <= DAYS_BEFORE_DEADLINE)
        ]
    
        if pending.empty:
            try:
                r = requests.post(WEBHOOK_URL, json={"content": "🎉 締切の近い、または過ぎた作業は現在ありません。"})
                write_log(f"No pending tasks. webhook status: {getattr(r,'status_code', 'N/A')}")
            except Exception as e:
                write_log("Webhook send failed (no pending): " + repr(e))
            return 0
    
        # Embedの色分け
        STYLE_MAP = {
            "デザイナー": {"color": 0xFFD700},	#黄色
            "プログラマー": {"color": 0x1E90FF},	#青
            "サウンド": {"color": 0xFFA500},		#オレンジ
            "未設定": {"color": 0x808080},		#グレー
        }

		# 折り返す文字数
        WRAP_RULES = {
            "職種": 8,
            "分類": 10,
            "内容": 14,
            "詳細": 34,
            "担当": 4,
            "進捗": 4,
            "優先度": 4,
            "備考": 10,
            "締切": 6,
        }
    
        def make_task_image(name, tasks, rate):
            DISPLAY_COLUMNS = [c for c in COLUMN_ORDER if c in COL_WIDTH_MAP]
            headers = DISPLAY_COLUMNS
        
            font_path = os.path.join(os.environ["WINDIR"], "Fonts", "meiryo.ttc")
        
            try:
                title_font = ImageFont.truetype(font_path, pt_to_px(16))
                header_font = ImageFont.truetype(font_path, pt_to_px(11))
                text_font = ImageFont.truetype(font_path, pt_to_px(11))
            except Exception:
                title_font = ImageFont.load_default()
                header_font = ImageFont.load_default()
                text_font = ImageFont.load_default()
        
            # 固定値・定義は先に
            TOP_PADDING = 6
            LEFT_PADDING = 10
            LEFT_ALIGN_COLUMNS = ["詳細", "備考"]
            line_height = 20
            MAX_HEIGHT = 5000
        
            STATUS_COLOR_MAP = {
                "完了": (180, 210, 255),
                "確認待ち": (180, 240, 200),
                "進行中": (255, 245, 170),
                "未着手": (245, 245, 245),
            }
        
            col_widths = [COL_WIDTH_MAP[h] for h in headers]
        
            # テキストを幅で折り返す
            def wrap_text_pixel(text, max_width):
                if not text:
                    return [""]
                dummy_img = Image.new("RGB", (1, 1))
                draw_dummy = ImageDraw.Draw(dummy_img)
                lines = []
                for raw_line in str(text).splitlines():
                    current = ""
                    for char in raw_line:
                        if draw_dummy.textlength(current + char, font=text_font) <= max_width - 12:
                            current += char
                        else:
                            lines.append(current)
                            current = char
                    lines.append(current)
                return lines
        
            wrapped_rows = []
            for _, row in tasks.iterrows():
                dl = row["締切"]
                try:
                    deadline_date = dl.date() if hasattr(dl, "date") else pd.to_datetime(dl).date()
                except Exception:
                    deadline_date = datetime.now().date()
                deadline_text = deadline_date.strftime("%m/%d")
        
                values = []
                for h in headers:
                    if h == "締切":
                        values.append(deadline_text)
                    elif h == "進捗":
                        status = str(row.get("進捗_raw", "")).strip()
                        status_icon_map = {"完了": "完了", "確認待ち": "確認待ち", "進行中": "進行中", "未着手": "未着手"}
                        values.append(status_icon_map.get(status, status))
                    else:
                        values.append(row.get(h, ""))
        
                wrapped = [wrap_text_pixel(val, col_widths[i]) for i, val in enumerate(values)]
                max_lines = max(len(cell) for cell in wrapped)
                status_raw = str(row.get("進捗_raw", "")).strip()
                wrapped_rows.append((wrapped, max_lines, status_raw))
        
            # 画像サイズ計算
            header_height = 140
            total_height = header_height + sum((max_lines * line_height + TOP_PADDING*2) for _, max_lines, _ in wrapped_rows) + 40 + 80
            total_height = min(total_height, MAX_HEIGHT)
            total_width = sum(col_widths) + 40
        
            img = Image.new("RGB", (total_width, total_height), "white")
            draw = ImageDraw.Draw(img)
        
            title = f"{name} の締切が近い、または過ぎた作業（完了率 {rate}%）"
            try:
                title_w = draw.textbbox((0,0), title, font=title_font)[2]
            except Exception:
                title_w = draw.textlength(title, font=title_font)
            draw.text(((total_width - title_w)/2, 20), title, fill="black", font=title_font)
        
            # ヘッダー描画
            y = 90
            x_start = 20
            x = x_start
            for i, header in enumerate(headers):
                draw.rectangle([x, y, x + col_widths[i], y + 45], fill=(230,230,230), outline="black", width=1)
                text_w = draw.textlength(header, font=header_font)
                draw.text((x + (col_widths[i]-text_w)/2, y+10), header, fill="black", font=header_font)
                x += col_widths[i]
            y += 45
        
            # 行描画
            for wrapped, max_lines, status_raw in wrapped_rows:
                row_height = max_lines * line_height + TOP_PADDING*2
                x = x_start
                bg_color = STATUS_COLOR_MAP.get(status_raw, (255,255,255))
                for col_index, (col_name, cell_lines) in enumerate(zip(headers, wrapped)):
                    w = col_widths[col_index]
                    draw.rectangle([x, y, x + w, y + row_height], fill=bg_color, outline="black", width=1)
                    total_cell_height = len(cell_lines)*line_height
                    if col_name in LEFT_ALIGN_COLUMNS:
                        start_y = y + TOP_PADDING
                    else:
                        start_y = y + (row_height - total_cell_height)/2
                    for i, line in enumerate(cell_lines):
                        line_y = start_y + i*line_height
                        if col_name in LEFT_ALIGN_COLUMNS:
                            draw.text((x + LEFT_PADDING, line_y), line, font=text_font, fill="black")
                        else:
                            line_w = draw.textlength(line, font=text_font)
                            draw.text((x + (w - line_w)/2, line_y), line, font=text_font, fill="black")
                    x += w
                y += row_height
        
            # バッファに保存
            buffer = BytesIO()
            img.save(buffer, format="PNG")
            buffer.seek(0)
            return buffer
        
        # 個別送信
        all_embeds = []
        errors = []
        
        for 担当表示, group in pending.groupby("担当"):
        
            namekey = str(担当表示).strip() if 担当表示 is not None else "未設定"
            rate = person_rates.get(namekey, 0)
        
            mention = ""
            if MENTION_ENABLED and isinstance(MENTION_MAP, dict):
                mention = MENTION_MAP.get(namekey, "")
            else:
                mention = ""
        
            try:
                image_buffer = make_task_image(namekey, group, rate)
        
                r = requests.post(
                    WEBHOOK_URL,
                    files={"file": ("task.png", image_buffer, "image/png")},
                    data={"content": f"📗 {mention} {namekey} の作業リスト"}
                )
                write_log(f"Posted image for {namekey}, status={getattr(r,'status_code','N/A')}")
        
            except Exception as e:
                write_log(f"Failed to post image for {namekey}: {repr(e)}")
                errors.append((namekey, repr(e)))
        
            representative_job = str(group.iloc[0].get("職種", "未設定")).strip()
            embed_color = STYLE_MAP.get(representative_job, STYLE_MAP["未設定"])["color"]
        
            # 担当者ごとのEmbed
            lines = []
        
            for _, row in group.iterrows():
                try:
                    days_left = (row["締切"].date() - today).days
                except Exception:
                    days_left = 0
        
                if days_left < 0:
                    days_text = f"🔴 締切が {abs(days_left)} 日過ぎてる！"
                elif days_left == 0:
                    days_text = "🟠 今日が締切！"
                elif days_left == 1:
                    days_text = "🟡 明日が締切！"
                elif days_left <= 3:
                    days_text = f"🟡 締切まであと {days_left} 日！"
                else:
                    days_text = f"⚪ 締切まであと {days_left} 日"
        
                lines.append(f"・{row['内容']}（{days_text}）")
        
            description_text = "\n".join(lines)
        
            if len(description_text) > 4000:
                description_text = description_text[:3900] + "\n…（以下省略）"
        
            embed = {
                "title": f"📋 {namekey}の作業一覧（完了率 {rate}%）",
                "description": description_text,
                "color": embed_color,
                "footer": {"text": f"更新日: {today.strftime('%Y/%m/%d')}"}
            }
        
            all_embeds.append(embed)
        
        # for の外に戻る
        
        all_embeds = all_embeds[:10]

        description_text = "\n".join(lines)
        
        if len(description_text) > 4000:
            description_text = description_text[:3900] + "\n…（以下省略）"
        
        if DAYS_BEFORE_DEADLINE == 0:
            deadline_text = "今日が締切の作業、または締切を過ぎた作業があるぞ！"
        elif DAYS_BEFORE_DEADLINE == 1:
            deadline_text = "明日が締切の作業、または締切を過ぎた作業があるぞ！"
        else:
            deadline_text = f"{DAYS_BEFORE_DEADLINE}日以内に締切の作業、または締切を過ぎた作業があるぞ！"
		
        payload = {
            "content": f"⚠️ **本日の締切連絡** ⚠️\n"
                       f"✅ 全体完了率：{overall_rate}%\n"
                       f"{deadline_text}",
            "embeds": all_embeds,
        }
    
        try:
            r2 = requests.post(WEBHOOK_URL, json=payload)
            write_log(f"Posted summary payload status={getattr(r2,'status_code','N/A')}")
            if r2.status_code >= 400:
                write_log("Summary post failed: " + (r2.text if hasattr(r2, "text") else ""))
        except Exception as e:
            write_log("Failed to post summary payload: " + repr(e))
    
        if errors:
            write_log("Some posting errors: " + repr(errors))
    
        write_log("正常終了")
        return 0
    
    except Exception:
        # エラーを記録
        try:
            with open(ERR_FILE, "w", encoding="utf-8") as f:
                f.write(traceback.format_exc())
        except Exception:
            pass
        write_log("EXCEPTION: see " + ERR_FILE)

        return 1
